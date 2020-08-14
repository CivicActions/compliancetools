#!/usr/bin/env python

import openpyxl
import yaml
from pathlib import Path
from shutil import copyfile

def main():
    config = load_options()
    if file_exists(config['original_file']):
        dir_exists(config['copy_to_directory'])
        original = Path(config['original_file'])
        out = Path(config['copy_to_directory']).joinpath(original.name)
        # Copy original spreadsheet file to path set in config.
        copyfile(original, out)

        sheet = config['sheet_name']
        start_row = config['starting_row']
        status_col = config['status_column']
        type_col = config['control_type_column']
        imp_col = config['implementation_column']
        control_col = config['control_column']

        # Get the data from the component files sorted by control key
        cd = get_control_data()
        # Open the spreadsheet and write the values.
        wb = openpyxl.load_workbook(filename=out, data_only=True)
        ws = wb[sheet]
        for key, row in enumerate(ws.iter_rows(min_row=start_row)):
            row_num = key + start_row
            if row[control_col].value is None:
                break
            ctrl = row[control_col].value
            status_key = ctrl.replace('(', '_(').replace('-0', '-')
            if status_key in cd:
                print("Updating {}".format(status_key))
                if 'status' in cd[status_key]:
                    ws[status_col+str(row_num)] = cd[status_key]['status']
                ws[type_col+str(row_num)] = cd[status_key]['type']
                narrative = get_narratives(cd[status_key]['narrative'])
                ws[imp_col+str(row_num)] = narrative
            else:
                ws[status_col+str(row_num)] = 'not applicable'
        wb.save(out)

def get_narratives(narr):
    '''Return the control narratives as a string.'''
    contractor = get_key_values('contractor.yaml', 'name_short')
    project = get_key_values('project.yaml', 'client_name_short')
    text = ''
    for k, n in narr.items():
        if k != 'no key':
            text = "{}Part {})".format(text, k.upper())
        for p in n:
            if p['parent'] == 'Project':
                parent = project
            elif p['parent'] == 'Contractor':
                parent = contractor
            else:
                parent = p['parent']
            text = "{}\n\r{} responsibility\n\r{}\n\r".format(text, parent, p['summary'])
    return text

def load_options():
    '''Load the spreadsheet options.'''
    options_file = Path('keys/spreadsheet.yaml')
    if options_file.is_file():
        with open(options_file, newline="") as stream:
            options = yaml.safe_load(stream)
            return options
    else:
        create_options()

def file_exists(file):
    '''Check that a file exists.'''
    try:
        Path(file).exists()
    except FileNotFoundError as e:
        print(e)
    return True

def dir_exists(dir):
    '''Check that a directory exists. Create it if it doesn't.'''
    newdir = Path(dir)
    if not newdir.exists():
        newdir.mkdir(parents=True)

def create_options():
    '''Create the spreadsheet.yaml file if one doesn't exist.'''
    opts = {
        'original_file': '',
        'copy_to_directory': '',
        'sheet_name': '',
        'control_type_column': '',
        'control_column': '6 # Column which contains the Control ID. For example: AC-2. Oddly, this needs to be a number not an alpha.',
        'status_column': '',
        'implementation_column': '',
        'starting_row': 1
    }
    dir_exists('keys')

    with open(Path('keys/spreadsheet.yaml'), mode='w', newline='\r\n') as opt_file:
        print('Creating keys/spreadsheet.yaml file.')
        yaml.dump(opts, opt_file)
    raise Exception('The values in the spreadsheet.yaml file are required.')

def get_control_data():
    oc = get_opencontrol()
    component_files = []
    # Create a list of all of the components in the project.
    for component in oc['components']:
        p = Path(component).rglob("*")
        files = [x for x in p if x.is_file() and x.name != 'component.yaml']
        component_files.extend(files)
        component_files.sort()
    controls = parse_controls(component_files)
    return controls

def get_opencontrol():
    '''Load the OpenControl.yaml file which points to the components
    that are used for creating the SSP.'''
    if file_exists('opencontol.yaml'):
        oc_yaml = Path('opencontrol.yaml')

    try:
        with open(oc_yaml, encoding="utf8") as f:
            try:
                opencontrol = yaml.safe_load(f)
            except Exception as e:
                raise ValueError("OpenControl file {} has invalid data (is not valid YAML: {}).".format(
                    oc_yaml,
                    str(e) ))
            if not isinstance(opencontrol, dict):
                raise ValueError("OpenControl file {} has invalid data (should be a mapping, is a {}).".format(
                    oc_yaml,
                    type(opencontrol)))
    except IOError as e:
        raise ValueError("OpenControl file {} could not be loaded: {}.".format(
            oc_yaml,
            str(e) ))
    return opencontrol

def parse_controls(components):
    '''Sort all of the controls by Control ID and the narratives
    by key.
    For example:
    'AC-2': {
        'type': 'Hybrid',
        'narrative': [{
            'a': [{
                'parent': 'AWS',
                'summary': 'AWS does this',
            }, {
                'parent': 'Drupal',
                'summary': 'Drupal does this
            }]
        }]
    }
    ...
    '''
    st = get_statuses()
    cd = {}
    for c in components:
        with open(c, newline='') as ct:
            control = yaml.safe_load(ct)
            for s in control['satisfies']:
                ctrl = str(s['control_key']).replace(' ', '_')
                if ctrl not in cd:
                    cd[ctrl] = {}
                cd[ctrl]['type'] = s['security_control_type']

                if ctrl in st:
                    cd[ctrl]['status'] = st[ctrl]

                for n in s['narrative']:
                    if 'narrative' not in cd[ctrl]:
                        cd[ctrl]['narrative'] = {}
                    sum = {'parent': c.parent.name, 'summary': n['text']}
                    if 'key' in n:
                        key = n['key']
                    else:
                        key = 'no key'

                    if key not in cd[ctrl]['narrative']:
                        cd[ctrl]['narrative'][key] = []
                    cd[ctrl]['narrative'][key].append(sum)
    return cd

def get_statuses():
    '''Load the status.yaml file.'''
    st = Path('keys/status.yaml')
    if st.exists:
        with open(st, "r", newline="") as stream:
            try:
                stat = yaml.safe_load(stream)
            except yaml.YAMLError as exc:
                print(exc)
        return stat

def get_key_values(file, key):
    '''Given a YAML file and a key, load the file and return the
    value for the given key.'''
    k = Path('keys').joinpath(file)
    if k.exists:
        with open(k, newline="") as stream:
            options = yaml.safe_load(stream)
    return options[key]

if __name__ == "__main__":
    main()